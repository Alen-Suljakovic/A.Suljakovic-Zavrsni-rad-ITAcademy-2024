from selenium import webdriver 
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import datetime

dateAndTime = datetime.datetime.now()
dateAndTime = dateAndTime.strftime("%d-%m-%Y %H-%M-%S")
rowNum = 3                                                     #Starts at 2 since the 1st excel Cell is for the title
i=2                                                 #Has start at 2 since no1 is for the title
x = None
lookupValue = None
cellValue = None
exception = 0
QACheckbook = "QA Checkbook " + str(dateAndTime) + ".xlsx"
centerText = Alignment(horizontal='center', vertical='center')              #Center text in excel
cellfillRed = PatternFill(patternType='solid', fgColor='FF3333')            #choose fill type and RED color in excel
cellfillGreen = PatternFill(patternType='solid', fgColor='33CC33')       #choose fill type and GREEN color

wb = openpyxl.load_workbook("QA Checkbook Template.xlsx")
ws = wb['Background data'] 
backgroundData = ws['B1'].value
wb = openpyxl.load_workbook(backgroundData)
QACheckbook = backgroundData

def wait(x):
    driver.implicitly_wait(x)                #Time module discarded- code runs faster

def findAndSendByName(Name,Input):           #Find an element by name and send input
    x = driver.find_element(By.NAME, Name)
    x.send_keys(Input)
    return x

def findAndSendByXPath(XPath,Input):
    x = driver.find_element(By.XPATH, XPath)
    x.send_keys(Input)
    return x

def findByName(Name):                        #Find an element by name
    return driver.find_element(By.NAME, Name)

def findByXPath(XPath):
    return driver.find_element(By.XPATH, XPath)

def findByCSSSelector(Selector):
    return driver.find_element(By.CSS_SELECTOR, Selector)

def testFailed():
    global rowNum
    cell = "E"
    cell += str(rowNum)                                  #Casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['2. Header and Footer']                              #Select excel sheesheet and load to ws
    ws[cell] = "FAIL"                            #Write to cell 
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillRed
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def testPassed():
    global rowNum
    cell = "E"
    cell += str(rowNum)                                  # casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['2. Header and Footer']                              #Select excel sheesheet and load to ws
    ws[cell] = "PASS"                            #Write to cell  
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillGreen
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def DoesNotExist():
    global rowNum
    print('DOES NOT EXIST!')
    cell = "E"
    cell += str(rowNum)                                  #Casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['2. Header and Footer']                              #Select excel sheesheet and load to ws
    ws[cell] = "FAIL"                            #Write to cell 
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillRed
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)
    
driver = webdriver.Chrome()

driver.get("https://puppies-closet.com/evidencija/login.php")

driver.maximize_window()


'''------------------------ LOGIN ----------------------------'''

username = findAndSendByName("username", "alensuljakovic")
password = findAndSendByName("password", "alensuljakovic123")

login = findByName("login")
login.click()

'''--------------------------------------TESTING HEADERS-----------------------------------------------------'''


wait(5)
landingPage = driver.title
findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(1) > a').click() #Check if 'Zaposleni - Zaduživanje/Razduživanje' works
if findByCSSSelector('#wrapper > main > section.section-two > p').text == 'LISTA ZAPOSLENIH':
    testPassed()
else:
    testFailed()

wait(5)
findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(2) > a').click() #Check if 'Oprema' works
if findByCSSSelector('#wrapper > main > section.section-two > p').text == 'LISTA OPREME':
    testPassed()
else:
    testFailed()

wait(5)
findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(3) > a').click() #Check if 'Izvjestaji' works
if findByCSSSelector('#wrapper > div > div:nth-child(1) > form > p').text == 'IZVJEŠTAJI O ZADUŽENOJ OPREMI PO ORGANIZACIONIM JEDINICAMA':
    testPassed()
else:
    testFailed()

wait(5)
findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(4) > a').click() #Check if 'tip/proizcodjas opreme' works
if findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(1) > label').text == 'PREGLED, IZMJENA, UNOS I BRISANJE TIPA OPREME':
    testPassed()
else:
    testFailed()

wait(5)
findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(5) > a').click() #Check if 'kancelarija/org jedinica' works
if findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(1) > label').text == 'PREGLED, IZMJENA, UNOS I BRISANJE BROJA KANCELARIJE':
    testPassed()
else:
    testFailed()

wait(5)
findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(6) > a').click() #Check if 'administracija korisnika' works
if findByCSSSelector('#wrapper > div.section-two > p').text == 'LISTA KORISNIKA':
    testPassed()
else:
    testFailed()

wait(5)
findByCSSSelector('#wrapper > header > nav > ul.logout > li > a').click() #Check if 'odjava' works
newPage = driver.title
if landingPage != newPage:
    testPassed()
else:
    testFailed()

'''--------------------------------------TESTING FOOTERS-----------------------------------------------------'''
username = findAndSendByName("username", "alensuljakovic")
password = findAndSendByName("password", "alensuljakovic123")
login = findByName("login")
login.click()

findByCSSSelector('#wrapper > footer > div.top-footer > div:nth-child(1) > a').click() #Check if 'Uputstvo za korištenje Elektronske evidencije računarske opreme' works
newPage = driver.title
if landingPage != newPage:
    testPassed()
else:
    testFailed()

findByCSSSelector('#wrapper > footer > div.top-footer > div:nth-child(2) > a').click() #Check if 'Uputstvo za korisnike za prijavu poteškoća u radu' works
newPage = driver.title
if landingPage != newPage:
    testPassed()
else:
    testFailed()

driver.quit()