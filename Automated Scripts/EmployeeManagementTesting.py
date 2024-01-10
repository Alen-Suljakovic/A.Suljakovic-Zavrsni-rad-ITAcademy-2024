from selenium import webdriver 
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import datetime

dateAndTime = datetime.datetime.now()
dateAndTime = dateAndTime.strftime("%d-%m-%Y %H-%M-%S")
rowNum = 3                                                     #Starts at 2 since the 1st excel Cell is for the title
i=2                                                 #Has start at 2 since no1 is for the title
x = None
lookupValue = None
cellValue = None
exception = 0
locatedNameXpath = 'PlaceHolder'
QACheckbook = "QA Checkbook " + str(dateAndTime) + ".xlsx"
centerText = Alignment(horizontal='center', vertical='center')              #Center text in excel
cellfillRed = PatternFill(patternType='solid', fgColor='FF3333')            #choose fill type and RED color in excel
cellfillGreen = PatternFill(patternType='solid', fgColor='33CC33')       #choose fill type and GREEN color

wb = openpyxl.load_workbook("QA Checkbook Template.xlsx")
ws = wb['Background data'] 
backgroundData = ws['B1'].value
wb = openpyxl.load_workbook(backgroundData)
QACheckbook = backgroundData

def wait(x):
    driver.implicitly_wait(x)                #Time module discarded- code runs faster

def findAndSendByName(Name,Input):           #Find an element by name and send input
    x = driver.find_element(By.NAME, Name)
    x.send_keys(Input)
    return x

def findAndSendByXPath(XPath,Input):
    x = driver.find_element(By.XPATH, XPath)
    x.send_keys(Input)
    return x

def findByName(Name):                        #Find an element by name
    return driver.find_element(By.NAME, Name)

def findByXPath(XPath):
    return driver.find_element(By.XPATH, XPath)

def findByCSSSelector(Selector):
    return driver.find_element(By.CSS_SELECTOR, Selector)

def findByCSSSelector2(Selector):  
    return driver2.find_element(By.CSS_SELECTOR, Selector)

def findAndSendByCSSSelector(Selector, Input):
        x = driver.find_element(By.CSS_SELECTOR, Selector)
        x.send_keys(Input)
        return x

def testFailed():
    global rowNum
    cell = "E"
    cell += str(rowNum)                                  #Casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['3. Employee management']                              #Select excel sheesheet and load to ws
    ws[cell] = "FAIL"                            #Write to cell 
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillRed
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def testPassed():
    global rowNum
    cell = "E"
    cell += str(rowNum)                                  # casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['3. Employee management']                              #Select excel sheesheet and load to ws
    ws[cell] = "PASS"                            #Write to cell  
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillGreen
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def DoesNotExist():
    global rowNum
    print('DOES NOT EXIST!')
    cell = "E"
    cell += str(rowNum)                                  #Casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['3. Employee management']                              #Select excel sheesheet and load to ws
    ws[cell] = "FAIL"                            #Write to cell 
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillRed
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def searchThroughTableforName(lookupValue):             #Search through table for Name Value
    global cellValue
    global exception
    global locatedNameXpath
    global i
    cellValue = "Reset"
    i = 2
    exception = 0
    while cellValue != lookupValue and exception != 1:
        try:
            locatedNameXpath = driver.find_element(By.CSS_SELECTOR, '#results > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(2)')
        except: 
            exception = 1
            print(lookupValue, " NOT found on list!")
            print(cellValue, " je cellValue")

        try:
            cellValue = locatedNameXpath.text
        except:
            cellValue = 'Blank'
        
        if cellValue == lookupValue:
            print(cellValue, 'Found in table!')
        elif cellValue != lookupValue:
            i +=1

def searchThroughTableforLastname(lookupValue):
    global cellValue
    global exception
    global locatedNameXpath
    global i
    cellValue = "Reset"
    i = 2
    exception = 0
    while cellValue != lookupValue and exception != 1:
        try:
            locatedNameXpath = driver.find_element(By.CSS_SELECTOR, '#results > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(3)')
        except: 
            exception = 1
            print("Lastname NOT found on list!")

        try:
            cellValue = locatedNameXpath.text
        except:
            cellValue = 'Blank'
        
        if cellValue == lookupValue:
            print(cellValue, 'Found in table!')
        elif cellValue != lookupValue:
            i +=1
    
def searchThroughTableforEmail(lookupValue):
    global cellValue
    global exception
    global locatedNameXpath
    global i
    cellValue = "Reset"
    i = 2
    exception = 0
    while cellValue != lookupValue and exception != 1:
        try:
            locatedNameXpath = driver.find_element(By.CSS_SELECTOR, '#results > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(4)')
        except: 
            exception = 1
            print("Email NOT found on list!")

        try:
            cellValue = locatedNameXpath.text
        except:
            cellValue = 'Blank'
        
        if cellValue == lookupValue:
            print(cellValue, 'Found in table!')
        elif cellValue != lookupValue:
            i +=1

def searchThroughTableforPhone(lookupValue):
    global cellValue
    global exception
    global locatedNameXpath
    global i
    cellValue = "Reset"
    i = 2
    exception = 0
    while cellValue != lookupValue and exception != 1:
        try:
            locatedNameXpath = driver.find_element(By.CSS_SELECTOR, '#results > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(5)')
        except: 
            exception = 1
            print("Phone NOT found on list!")

        try:
            cellValue = locatedNameXpath.text
        except:
            cellValue = 'Blank'
        
        if cellValue == lookupValue:
            print(cellValue, 'Found in table!')
        elif cellValue != lookupValue:
            i +=1

def searchThroughTableforoffice(lookupValue):            #Search through table for Name Value
    global cellValue
    global exception
    global locatedNameXpath
    global i
    cellValue = "Reset"
    i = 2
    exception = 0
    while cellValue != lookupValue and exception != 1:
        try:
            locatedNameXpath = driver.find_element(By.CSS_SELECTOR, '#results > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(6)')
        except: 
            exception = 1
            print("Office NOT found on list!")

        try:
            cellValue = locatedNameXpath.text
        except:
            cellValue = 'Blank'
        
        if cellValue == lookupValue:
            print(cellValue, 'Found in table!')
        elif cellValue != lookupValue:
            i +=1

def searchThroughTableforOrgUnit(lookupValue):            #Search through table for Name Value
    global cellValue
    global exception
    global locatedNameXpath
    global i
    cellValue = "Reset"
    i = 2
    exception = 0
    while cellValue != lookupValue and exception != 1:
        try:
            locatedNameXpath = driver.find_element(By.CSS_SELECTOR, '#results > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(7)')
        except: 
            exception = 1
            print("OrgUnit NOT found on list!")

        try:
            cellValue = locatedNameXpath.text
        except:
            cellValue = 'Blank'
        
        if cellValue == lookupValue:
            print(cellValue, 'Found in table!')
        elif cellValue != lookupValue:
            i +=1

driver = webdriver.Chrome()

driver.get("https://puppies-closet.com/evidencija/login.php")

driver.maximize_window()


'''------------------------ LOGIN ----------------------------'''

username = findAndSendByName("username", "alensuljakovic")
password = findAndSendByName("password", "alensuljakovic123")

title = driver.title

login = findByName("login")
login.click()

'''----------------------------------------ADDING NEW ACCOUNT----------------------------------------------'''

findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(1) > a').click()

findAndSendByName('firstname', 'AlenTest1')
findAndSendByName('lastname', 'SuljakovicTest1')
findAndSendByName('email', 'AlenDummy@QA.com')
findAndSendByName('phone', '111111111')
findByName('office_id').click()
findByCSSSelector('#office_id > option:nth-child(14)').click()
findByName('organization_id').click()
findByCSSSelector('#organization_id > option:nth-child(18)').click()

time.sleep(5)
findByCSSSelector('#wrapper > main > section.section-one > div.section-one-left > form > input.button.blue').click()
time.sleep(5)

findAndSendByName('search', 'AlenTest1')
findByName('employeesSearch').click()
time.sleep(3)
searchThroughTableforName('AlenTest1')

if cellValue == 'AlenTest1':
    testPassed()
else:
    testFailed()
print('I je:', i)
time.sleep(3)

'''----------------------------------------UPDATING ACCOUNT (everything at once)----------------------------------------------'''

findByCSSSelector('#results > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(8) > button:nth-child(1)').click()

time.sleep(3)


newName = 'AlenUpdated'
newLastname = 'SuljakovicUpdated'
newEmail = 'EmailUpdated@QA.com'
newPhone = '222222222'

findByCSSSelector('#modalemployee > div > div.modal-body > div > form > input[type=text]:nth-child(1)').clear()
findAndSendByCSSSelector('#modalemployee > div > div.modal-body > div > form > input[type=text]:nth-child(1)', newName) #update name

findByCSSSelector('#modalemployee > div > div.modal-body > div > form > input[type=text]:nth-child(2)').clear()
findAndSendByCSSSelector('#modalemployee > div > div.modal-body > div > form > input[type=text]:nth-child(2)', newLastname) #update name

findByCSSSelector('#modalemployee > div > div.modal-body > div > form > input[type=text]:nth-child(3)').clear()
findAndSendByCSSSelector('#modalemployee > div > div.modal-body > div > form > input[type=text]:nth-child(3)', newEmail) #update name

findByCSSSelector('#modalemployee > div > div.modal-body > div > form > input[type=text]:nth-child(4)').clear()
findAndSendByCSSSelector('#modalemployee > div > div.modal-body > div > form > input[type=text]:nth-child(4)', newPhone) #update name

driver2 = findByCSSSelector('#modalemployee > div > div.modal-body') #loading in the window prompt a

findByCSSSelector2('#office_id').click()
findByCSSSelector2('#office_id > option:nth-child(29)').click()

findByCSSSelector2('#organization_id').click()
findByCSSSelector2('#organization_id > option:nth-child(11)').click()

findByCSSSelector('#modalemployee > div > div.modal-body > div > form > input.button.blue').click()

time.sleep(3)
findAndSendByName('search', newName)
findByName('employeesSearch').click()
time.sleep(3)

searchThroughTableforName(newName)
if cellValue == newName:
    testPassed()
else:
    testFailed()

searchThroughTableforLastname(newLastname)
if cellValue == newLastname:
    testPassed()
else:
    testFailed()

searchThroughTableforEmail(newEmail)
if cellValue == newEmail:
    testPassed()
else:
    testFailed()

searchThroughTableforPhone(newPhone)
if cellValue == newPhone:
    testPassed()
else:
    testFailed()

newoffice = findByCSSSelector('#results > div > table > tbody > tr:nth-child(2) > td:nth-child(6)').text

if newoffice != '1014':
    testPassed()
else:
    testFailed()

newOrgUnit = findByCSSSelector('#results > div > table > tbody > tr:nth-child(2) > td:nth-child(7)').text

if newoffice != 'plava':
    testPassed()
else:
    testFailed()


'''----------------------------------------DELETING ACCOUNT----------------------------------------------'''

searchThroughTableforName(newName)

findByCSSSelector('#results > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(8) > button.button.red').click()
findByCSSSelector('#del').click()

time.sleep(3)
findAndSendByName('search', newName)
findByName('employeesSearch').click()
time.sleep(3)

searchThroughTableforName(newName)

if cellValue == newName:
    testPassed()
else:
    testFailed()


