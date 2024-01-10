from selenium import webdriver 
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import datetime

dateAndTime = datetime.datetime.now()
dateAndTime = dateAndTime.strftime("%d-%m-%Y %H-%M-%S")
rowNum = 3                                                     #Starts at 2 since the 1st excel Cell is for the title
i=2                                                 #Has start at 2 since no1 is for the title
x = None
lookupValue = None
cellValue = None
exception = 0
QACheckbook = "QA Checkbook " + str(dateAndTime) + ".xlsx"
centerText = Alignment(horizontal='center', vertical='center')              #Center text in excel
cellfillRed = PatternFill(patternType='solid', fgColor='FF3333')            #choose fill type and RED color in excel
cellfillGreen = PatternFill(patternType='solid', fgColor='33CC33')       #choose fill type and GREEN color

wb = openpyxl.load_workbook("QA Checkbook Template.xlsx")
ws = wb['Background data'] 
ws['B1'] = QACheckbook
backgroundData = ws['B1'].value
wb.save("QA Checkbook Template.xlsx")
QACheckbook = backgroundData

def wait(x):
    driver.implicitly_wait(x)                #Time module discarded- code runs faster

def findAndSendByName(Name,Input):           #Find an element by name and send input
    x = driver.find_element(By.NAME, Name)
    x.send_keys(Input)
    return x

def findAndSendByXPath(XPath,Input):
    x = driver.find_element(By.XPATH, XPath)
    x.send_keys(Input)
    return x

def findByName(Name):                        #Find an element by name
    return driver.find_element(By.NAME, Name)

def findByXPath(XPath):
    return driver.find_element(By.XPATH, XPath)

def findByCSSSelector(Selector):
    return driver.find_element(By.CSS_SELECTOR, Selector)

def findByCSSSelector2(Selector):  
    return driver2.find_element(By.CSS_SELECTOR, Selector)

def findAndSendByCSSSelector(Selector, Input):
        x = driver.find_element(By.CSS_SELECTOR, Selector)
        x.send_keys(Input)
        return x

def testFailed():
    global rowNum
    cell = "E"
    cell += str(rowNum)                                  #Casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['1. Account management']                              #Select excel sheesheet and load to ws
    ws[cell] = "FAIL"                            #Write to cell 
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillRed
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def testPassed():
    global rowNum
    cell = "E"
    cell += str(rowNum)                                  # casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['1. Account management']                              #Select excel sheesheet and load to ws
    ws[cell] = "PASS"                            #Write to cell  
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillGreen
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def DoesNotExist():
    global rowNum
    print('DOES NOT EXIST!')
    cell = "E"
    cell += str(rowNum)                                  #Casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['1. Account management']                              #Select excel sheesheet and load to ws
    ws[cell] = "FAIL"                            #Write to cell 
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillRed
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def searchThroughTableforName(lookupValue):             #Search through table for Name Value
    global cellValue
    global exception
    global locatedNameXpath
    global i
    cellValue = "Reset"
    i = 3
    exception = 0
    while cellValue != lookupValue and exception != 1:
        try:
            locatedNameXpath = driver.find_element(By.CSS_SELECTOR, '#wrapper > div.section-two > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(2)') 
        except:
            exception = 1
            print("Name not found on list!")

        cellValue = locatedNameXpath.text

        if cellValue == lookupValue:
            print(cellValue, 'Found in table!')
        elif cellValue != lookupValue:
            i +=1

def searchThroughTableforLastName(lookupValue):             #Search through table for LastName Value
    global cellValue
    global exception
    global locatedNameXpath
    global i
    cellValue = "Reset"
    i = 3
    exception = 0
    while cellValue != lookupValue and exception != 1:
        try:
            locatedNameXpath = driver.find_element(By.CSS_SELECTOR, '#wrapper > div.section-two > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(3)')
        except:
            exception = 1
            print("LastName not found on list!")

        cellValue = locatedNameXpath.text

        if cellValue == lookupValue:
            print(cellValue, 'Found in table!')
        elif cellValue != lookupValue:
            i +=1

def searchThroughTableforUserName(lookupValue):             #Search through table for UserName Value
    global cellValue
    global exception
    global locatedNameXpath
    global i
    cellValue = "Reset"
    i = 3
    exception = 0
    while cellValue != lookupValue and exception != 1:
        try:
            locatedNameXpath = driver.find_element(By.CSS_SELECTOR, '#wrapper > div.section-two > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(4)')
        except:
            exception = 1
            print("UserName not found on list!")
        try:
            cellValue = locatedNameXpath.text
        except:
            print('Blank')
            cellValue = "blank"
        if cellValue == lookupValue:
            print(cellValue, 'Found in table!')
        elif cellValue != lookupValue:
            i +=1


driver = webdriver.Chrome()

driver.get("https://puppies-closet.com/evidencija/login.php")

driver.maximize_window()


'''------------------------TESTING LOGGIN IN WITH A WRONG PASSWORD----------------------------'''

username = findAndSendByName("username", "alensuljakovic")
password = findAndSendByName("password", "DummyWrongPass")

title = driver.title

login = findByName("login")
login.click()

expTitle = driver.title 

if title == expTitle:                        #Did we log in? if we did write FAILED in excel
    testPassed()
else:
    driver.back()
    testFailed()


'''-----------------------------TESTING LOGGIN IN WITH A CORRECT PASSWORD-----------------------------------'''

username = findAndSendByName("username", "alensuljakovic")
password = findAndSendByName("password", "alensuljakovic123")

title = driver.title

login = findByName("login")
login.click()

expTitle = driver.title 

if title != expTitle:                        #Did we log in? if not write FAILED in excel
    testPassed()
else:
    testFailed()

'''----------------------------------USER WITH STANDARD CHARACTERS----------------------------------------'''

userAdm = findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(6) > a')
userAdm.click()

userFirstName = 'AlenTest'                                              #USE VARS TO FEED VALUES
UserFirstNameObject = findAndSendByName('firstname', userFirstName)                 #Fills out the new user form !!Exception for faster code later
UserSurname = 'SuljakovicTest'
UserSurnameObject = findAndSendByName('lastname', UserSurname)
Username = 'AlenSuljak'
UsernameObject = findAndSendByName('username', Username)
UserPassword = 'alensuljak123'
UserPasswordObject = findAndSendByName('pass', UserPassword)

newUserAccType = findByName('role')
newUserAccType.click()
role = findByCSSSelector('#role > option:nth-child(3)')
role.click()


#  UNCOMMENT ONLY IF SURE TO CREATE NEW USER!
Submit = findByName('saveUser')
Submit.click()


searchThroughTableforName(userFirstName)

time.sleep(5)

if cellValue == userFirstName:
    testPassed()
else:
    testFailed()


'''-------------------USER WITH SPECIAL CHARACTERS-------------------'''

userFirstSpecName = 'AlenTest!'                                              #USE VARS TO FEED VALUES
UserFirstSpecNameObject = findAndSendByName('firstname', userFirstName)                 #Fills out the new user form !!Exception for faster code later
UserSpecSurname = 'SuljakovicTest!'
UserSpecSurnameObject = findAndSendByName('lastname', UserSurname)
SpecUsername = 'AlenSuljak!'
SpecUsernameObject = findAndSendByName('username', Username)
UserPassword = 'alensuljak123!'
UserPasswordObject = findAndSendByName('pass', UserPassword)

newUserAccType = findByName('role')
newUserAccType.click()
role = findByCSSSelector('#role > option:nth-child(3)')
accType = role.text
role.click()


Submit = findByName('saveUser')
Submit.click()


searchThroughTableforName(userFirstName)

if cellValue == userFirstName:
    testPassed()
else:
    testFailed()

'''----------------------------------------UPDATE NAME AND CHECK IF IT UPDATED-----------------------------'''
searchThroughTableforName(userFirstName)   

if exception != 1:
    updatedName= 'AlenUpdated'
    driver.find_element(By.CSS_SELECTOR, '#wrapper > div.section-two > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(6) > button:nth-child(1)').click() #click on edit button
    time.sleep(2)
    newFirstName = findByCSSSelector('#modaltype > div > div.modal-body > div > form > input[type=text]:nth-child(1)') 
    newFirstName.clear()
    newFirstName.send_keys(updatedName)
    findByCSSSelector('#modaltype > div > div.modal-body > div > form > input.button.blue').click()
    i=2
    searchThroughTableforName(updatedName)

    if cellValue == updatedName:
        testPassed()
    else:
        testFailed()
else:
    DoesNotExist()

time.sleep(5)




'''-----------------------------UPDATE LASTNAME AND CHECK IF UPDATED----------------------------------'''
i = 2 #redecare to reset looping
exception = 0
searchThroughTableforLastName(UserSurname)

if exception != 1:
    updatedSurname = 'SuljakovicUpdated'
    driver.find_element(By.CSS_SELECTOR, '#wrapper > div.section-two > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(6) > button:nth-child(1)').click() #click on edit button
    time.sleep(2)
    newSurname = findByCSSSelector('#modaltype > div > div.modal-body > div > form > input[type=text]:nth-child(2)')
    newSurname.clear()
    newSurname.send_keys(updatedSurname)
    findByCSSSelector('#modaltype > div > div.modal-body > div > form > input.button.blue').click()
    i = 2
    searchThroughTableforLastName(updatedSurname)
    print(cellValue)
    print(updatedSurname)
    if cellValue == updatedSurname:
        testPassed()
    else:
        testFailed()
else:
    DoesNotExist()

time.sleep(5)


'''-----------------------------UPDATE USERNAME AND CHECK IF UPDATED----------------------------------'''
i = 2 #redecalre i and exception to reset looping
exception = 0
searchThroughTableforUserName(Username)
if exception != 1:
    updatedUserName = 'AlenSuljakUpdated'
    driver.find_element(By.CSS_SELECTOR, '#wrapper > div.section-two > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(6) > button:nth-child(1)').click() #click on edit button
    time.sleep(2)
    newUsername = findByCSSSelector('#modaltype > div > div.modal-body > div > form > input[type=text]:nth-child(3)') 
    newUsername.clear()
    newUsername.send_keys(updatedUserName)
    findByCSSSelector('#modaltype > div > div.modal-body > div > form > input.button.blue').click() #click on submit
    i = 2
    searchThroughTableforUserName(updatedUserName)
    print(cellValue)
    if cellValue == updatedUserName:
        testPassed()
    else:
        testFailed()
else:
    DoesNotExist()

time.sleep(5)

'''-----------------------------UPDATE USER ACC TYPE AND CHECK IF UPDATED----------------------------------'''

'''-----------------------------CHOOSE WHAT ACCOUNT TO UPDATE!-------------------------------------------'''
accountToUpdate = 'AlenUpdated'
i = 2
exception = 0
searchThroughTableforName(accountToUpdate)
accountField = findByCSSSelector('#wrapper > div.section-two > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(5)') #reads accoutn type of found user
cellValue = accountField.text


if cellValue == 'Administrator':
    driver.find_element(By.CSS_SELECTOR, '#wrapper > div.section-two > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(6) > button:nth-child(1)').click()
    time.sleep(2)
    driver2 = findByCSSSelector('#modaltype > div')
    findByCSSSelector2('#role').click()
    findByCSSSelector2('#role > option:nth-child(1)').click()
    findByCSSSelector('#modaltype > div > div.modal-body > div > form > input.button.blue').click()
    i = 2 #redecalre i to reset looping
    searchThroughTableforName(accountToUpdate)
    accountField = findByCSSSelector('#wrapper > div.section-two > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(5)') #reads account type of found desired user
    accountFieldValue = accountField.text
    cellValue = accountFieldValue
    if cellValue == 'Korisnik':
        testPassed()
    else:
        testFailed()
else:
    driver.find_element(By.CSS_SELECTOR, '#wrapper > div.section-two > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(6) > button:nth-child(1)').click()
    time.sleep(2)
    driver2 = findByCSSSelector('#modaltype > div')
    findByCSSSelector2('#role').click()
    findByCSSSelector2('#role > option:nth-child(2)').click()
    findByCSSSelector('#modaltype > div > div.modal-body > div > form > input.button.blue').click()
    i = 2 #redecalre i to reset looping
    searchThroughTableforName(accountToUpdate)
    accountField = findByCSSSelector('#wrapper > div.section-two > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(5)') #reads account type of found desired user
    accountFieldValue = accountField.text
    cellValue = accountFieldValue
    if cellValue == 'Administrator':
        testPassed()
    else:
        testFailed()

'''---------------------------------UPDATING PASSWORD----------------------------------------'''
time.sleep(3)
searchThroughTableforName('AlenUpdated')
time.sleep(3)
findByCSSSelector('#wrapper > div.section-two > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(6) > button.button.blue').click()
time.sleep(3)
findByCSSSelector('#modaltype > div > div.modal-body > div > form > input[type=password]:nth-child(1)').clear()
findAndSendByCSSSelector('#modaltype > div > div.modal-body > div > form > input[type=password]:nth-child(1)', 'NewPassword')
findByCSSSelector('#modaltype > div > div.modal-body > div > form > input.button.blue').click()
findByCSSSelector('#wrapper > header > nav > ul.logout > li > a').click()

username = findAndSendByName("username", "AlenSuljakUpdated")
password = findAndSendByName("password", "NewPassword")

title = driver.title

login = findByName("login")
login.click()

expTitle = driver.title 

if title != expTitle:                        #Did we log in? if the page didn't change write FAILED in excel
    testPassed()
else:
    testFailed()

time.sleep(3)
findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(6) > a').click()
findByCSSSelector('#wrapper > div.section-two > div > table > tbody > tr:nth-child('+str(i)+') > td:nth-child(6) > button.button.red').click()
findByCSSSelector('#del').click()

searchThroughTableforUserName('alensuljakUpdated')

if exception == 1:
    testPassed()
else:
    testFailed()

time.sleep(10)

driver.quit()