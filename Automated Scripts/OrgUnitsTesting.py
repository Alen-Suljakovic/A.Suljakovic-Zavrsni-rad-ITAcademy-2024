from selenium import webdriver 
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import datetime

dateAndTime = datetime.datetime.now()
dateAndTime = dateAndTime.strftime("%d-%m-%Y %H-%M-%S")
rowNum = 3                                                     #Starts at 2 since the 1st excel Cell is for the title
i=2                                                 #Has start at 2 since no1 is for the title
x = None
lookupValue = None
cellValue = None
exception = 0
QACheckbook = "QA Checkbook " + str(dateAndTime) + ".xlsx"
centerText = Alignment(horizontal='center', vertical='center')              #Center text in excel
cellfillRed = PatternFill(patternType='solid', fgColor='FF3333')            #choose fill type and RED color in excel
cellfillGreen = PatternFill(patternType='solid', fgColor='33CC33')       #choose fill type and GREEN color
driver2 = None

wb = openpyxl.load_workbook("QA Checkbook Template.xlsx")
ws = wb['Background data'] 
backgroundData = ws['B1'].value
wb = openpyxl.load_workbook(backgroundData)
QACheckbook = backgroundData

def wait(x):
    driver.implicitly_wait(x)                #Time module discarded- code runs faster

def findAndSendByName(Name,Input):           #Find an element by name and send input
    try: 
        x = driver.find_element(By.NAME, Name)
        x.send_keys(Input)
        return x
    except:
        x = driver2.find_element(By.NAME, Name)
        x.send_keys(Input)
        return x

def findAndSendByName2(Name,Input):           #Find an element by name and send input
        x = driver2.find_element(By.NAME, Name)
        x.send_keys(Input)
        return x

def findByName(Name):                        #Find an element by name
        return driver.find_element(By.NAME, Name)
    
def findByName2(Name):                        #Find an element by name for org unit
        return driver2.find_element(By.NAME, Name)

def findByCSSSelector(Selector):
    return driver.find_element(By.CSS_SELECTOR, Selector)

def findByCSSSelector2(Selector):  
    return driver2.find_element(By.CSS_SELECTOR, Selector)

def findAndSendByCSSSelector(Selector, Input):
        x = driver.find_element(By.CSS_SELECTOR, Selector)
        x.send_keys(Input)
        return x

def findAndSendByCSSSelector2(Selector, Input):
    x = driver2.find_element(By.CSS_SELECTOR, Selector)
    x.send_keys(Input)
    return x

def testFailed():
    global rowNum
    cell = "E"
    cell += str(rowNum)                                  #Casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['7. Org and Office']                              #Select excel sheesheet and load to ws
    ws[cell] = "FAIL"                            #Write to cell 
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillRed
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def testPassed():
    global rowNum
    cell = "E"
    cell += str(rowNum)                                  # casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['7. Org and Office']                              #Select excel sheesheet and load to ws
    ws[cell] = "PASS"                            #Write to cell  
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillGreen
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def DoesNotExist():
    global rowNum
    print('DOES NOT EXIST!')
    cell = "E"
    cell += str(rowNum)                                  #Casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['7. Org and Office']                              #Select excel sheesheet and load to ws
    ws[cell] = "FAIL"                            #Write to cell 
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillRed
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def searchThroughTableforOffice(lookupValue):             #Search through table for UserName Value
    global cellValue
    global exception
    global i
    cellValue = "Reset"
    i = 3
    exception = 0
    while cellValue != lookupValue and exception != 1:
        try:
            locatedNameXpath = driver.find_element(By.CSS_SELECTOR, '#myTable > tbody > tr:nth-child('+str(i)+') > td:nth-child(2)')
        except:
            exception = 1
            print("Office not found on list!")

        cellValue = locatedNameXpath.text
        
        if cellValue == lookupValue:
            print(cellValue, 'Found in table at ', i)
        elif cellValue != lookupValue:
            i +=1

def searchThroughTableforOrgUnit(lookupValue):             #Search through table for UserName Value
    global cellValue
    global exception
    global i
    cellValue = "Reset"
    i = 3
    exception = 0
    while cellValue != lookupValue and exception != 1:
        try:
            tabela2 = findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > div > div:nth-child(2) > div.table.organization')
            locatedNameXpath = tabela2.find_element(By.CSS_SELECTOR, '#myTable > tbody > tr:nth-child('+str(i)+') > td:nth-child(2)')
        except:
            exception = 1
            print("Org Unit not found on list!")

        cellValue = locatedNameXpath.text
        
        if cellValue == lookupValue:
            print(cellValue, 'Found in table at ', i)
        elif cellValue != lookupValue:
            i +=1


driver = webdriver.Chrome()

driver.get("https://puppies-closet.com/evidencija/login.php")

driver.maximize_window()


'''------------------------ LOGIN ----------------------------'''

username = findAndSendByName("username", "alensuljakovic")
password = findAndSendByName("password", "alensuljakovic123")

title = driver.title

login = findByName("login")
login.click()

'''------------------------ OFFICE TESTING ----------------------------'''
'''------------------------ ADDING A NEW OFFICE ----------------------------'''

findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(5) > a').click() #go to 'kancelarija org. jed.'

findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(1) > label').click()
officeNum = "1999"
findAndSendByName('office', officeNum)
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(1) > div > div:nth-child(1) > form > input.button.blue').click()

'''------------------------ CHECK IF OFFICE ADDED ----------------------------'''

findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(1) > label').click()

searchThroughTableforOffice(officeNum)

if cellValue == officeNum: #Successfully added
    testPassed()
else:
    testFailed()


'''------------------------ UPDATING OFFICE ----------------------------'''
#we can stick with the same i value as the office will stay in the same place, otherwise add another 'searchThroughTableforOffice'
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(1) > label').click()
findByCSSSelector('#myTable > tbody > tr:nth-child('+str(i)+') > td:nth-child(3) > button:nth-child(1)').click()
time.sleep(3)
newOfficeNum = '2000'
findByCSSSelector('#modaloffice > div > div.modal-body > div > form > input[type=text]:nth-child(1)').clear()
findAndSendByCSSSelector('#modaloffice > div > div.modal-body > div > form > input[type=text]:nth-child(1)', newOfficeNum)
findByCSSSelector('#modaloffice > div > div.modal-body > div > form > input.button.blue').click()
time.sleep(3)
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(1) > label').click()
time.sleep(3)
searchThroughTableforOffice(newOfficeNum)

if cellValue == newOfficeNum: #testing if updated
    testPassed()
else:
    testFailed()

'''------------------------ DELETING OFFICE ----------------------------'''
time.sleep(3)
findByCSSSelector('#myTable > tbody > tr:nth-child('+str(i)+') > td:nth-child(3) > button.button.red').click()
findByCSSSelector('#del').click()
time.sleep(3)
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(1) > label').click()

searchThroughTableforOffice(newOfficeNum)

if cellValue != newOfficeNum: #we deleted it
    testPassed()
else:
    testFailed()



'''------------------------ ORG UNIT TESTING ----------------------------'''
'''------------------------ ADDING A NEW ORG UNITS ----------------------------'''



findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(5) > a').click() #go to 'kancelarija org. jed.'

findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > label').click()
orgUnit = "Odsjek 1999"
findAndSendByName('organization', orgUnit)
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > div > div:nth-child(1) > form > input.button.blue').click()

'''------------------------ CHECK IF ORH UNIT ADDED ----------------------------'''

findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > label').click()

searchThroughTableforOrgUnit(orgUnit)

if str(cellValue) == orgUnit: #Successfully added
    testPassed()
else:
    testFailed()


'''------------------------ UPDATING ORG UNIT ----------------------------'''
#we can stick with the same i value as the org unit will stay in the same place, otherwise add another 'searchThroughTableforOffice'
driver2 = findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > div') #load the table div to the driver variable so we can search through it, the 1st and 2nd table have the same selector
findByCSSSelector2('#myTable > tbody > tr:nth-child('+str(i)+') > td:nth-child(3) > button:nth-child(1)').click()

time.sleep(3)
newOrgUnitNum = 'Odsjek 2000'
findByCSSSelector('#modalorganization > div > div.modal-body > div > form > input[type=text]:nth-child(1)').clear()
findAndSendByCSSSelector('#modalorganization > div > div.modal-body > div > form > input[type=text]:nth-child(1)', newOrgUnitNum)
findByCSSSelector('#modalorganization > div > div.modal-body > div > form > input.button.blue').click()
time.sleep(3)
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > label').click()
searchThroughTableforOrgUnit(newOrgUnitNum)

if cellValue == newOrgUnitNum: #testing if updated
    testPassed()
else:
    testFailed()

'''------------------------ DELETING ORG UNIT ----------------------------'''

driver2 = findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > div')

time.sleep(15)
findByCSSSelector2('#myTable > tbody > tr:nth-child('+str(i)+') > td:nth-child(3) > button.button.red').click()
findByCSSSelector('#del').click()
time.sleep(3)
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > label').click()

searchThroughTableforOrgUnit(newOrgUnitNum)

if cellValue != newOrgUnitNum: #we deleted it 
    testPassed()
else:
    testFailed()


