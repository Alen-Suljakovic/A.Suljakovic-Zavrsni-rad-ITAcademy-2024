from selenium import webdriver 
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import datetime

dateAndTime = datetime.datetime.now()
dateAndTime = dateAndTime.strftime("%d-%m-%Y %H-%M-%S")
rowNum = 3                                                     #Starts at 2 since the 1st excel Cell is for the title
i=2                                                 #Has start at 2 since no1 is for the title
x = None
lookupValue = None
cellValue = None
exception = 0
locatedNameCSS = 'PlaceHolder'
QACheckbook = "QA Checkbook " + str(dateAndTime) + ".xlsx"
centerText = Alignment(horizontal='center', vertical='center')              #Center text in excel
cellfillRed = PatternFill(patternType='solid', fgColor='FF3333')            #choose fill type and RED color in excel
cellfillGreen = PatternFill(patternType='solid', fgColor='33CC33')       #choose fill type and GREEN color

wb = openpyxl.load_workbook("QA Checkbook Template.xlsx")
ws = wb['Background data'] 
backgroundData = ws['B1'].value
wb = openpyxl.load_workbook(backgroundData)
QACheckbook = backgroundData

def wait(x):
    driver.implicitly_wait(x)                #Time module discarded- code runs faster

def findAndSendByName(Name,Input):           #Find an element by name and send input
    x = driver.find_element(By.NAME, Name)
    x.send_keys(Input)
    return x

def findAndSendByXPath(XPath,Input):
    x = driver.find_element(By.XPATH, XPath)
    x.send_keys(Input)
    return x

def findByName(Name):                        #Find an element by name
    return driver.find_element(By.NAME, Name)

def findByXPath(XPath):
    return driver.find_element(By.XPATH, XPath)

def findByCSSSelector(Selector):
    return driver.find_element(By.CSS_SELECTOR, Selector)

def findByCSSSelector2(Selector):  
    return driver2.find_element(By.CSS_SELECTOR, Selector)

def findAndSendByCSSSelector(Selector, Input):
        x = driver.find_element(By.CSS_SELECTOR, Selector)
        x.send_keys(Input)
        return x

def testFailed():
    global rowNum
    cell = "E"
    cell += str(rowNum)                                  #Casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['5. Brand Testing']                              #Select excel sheesheet and load to ws
    ws[cell] = "FAIL"                            #Write to cell 
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillRed
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def testPassed():
    global rowNum
    cell = "E"
    cell += str(rowNum)                                  # casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['5. Brand Testing']                              #Select excel sheesheet and load to ws
    ws[cell] = "PASS"                            #Write to cell  
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillGreen
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def DoesNotExist():
    global rowNum
    print('DOES NOT EXIST!')
    cell = "E"
    cell += str(rowNum)                                  #Casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['5. Brand Testing']                              #Select excel sheesheet and load to ws
    ws[cell] = "FAIL"                            #Write to cell 
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillRed
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def searchThroughTableforEquipment(lookupValue):             #Search through table for Name Value
    global cellValue
    global exception
    global i
    cellValue = "Reset"
    i = 3
    exception = 0
    while cellValue != lookupValue and exception != 1:
        try:
            locatedNameXpath = driver.find_element(By.CSS_SELECTOR, '#myTable > tbody > tr:nth-child('+str(i)+') > td:nth-child(2)')
        except:
            exception = 1
            print("Equipment not found on list!")

        cellValue = locatedNameXpath.text
        
        if cellValue == lookupValue:
            print(cellValue, 'Found in table at ', i)
        elif cellValue != lookupValue:
            i +=1

def searchThroughTableforBrand(lookupValue):             #Search through table for Name Value+
    global cellValue
    global exception
    global i
    cellValue = "Reset"
    i = 3
    exception = 0
    while cellValue != lookupValue and exception != 1:
        try:
            tabela2 = findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > div > div:nth-child(2) > div.table.equipproducer')
            locatedNameXpath = tabela2.find_element(By.CSS_SELECTOR, '#myTable > tbody > tr:nth-child('+str(i)+') > td:nth-child(2)')
        except:
            exception = 1
            print("Brand not found on list!")

        cellValue = locatedNameXpath.text

        if cellValue == lookupValue:
            print(cellValue, 'Found in table!')
        elif cellValue != lookupValue:
            i +=1

driver = webdriver.Chrome()

driver.get("https://puppies-closet.com/evidencija/login.php")

driver.maximize_window()


'''------------------------LOGIN----------------------------'''

username = findAndSendByName("username", "alensuljakovic")
password = findAndSendByName("password", "alensuljakovic123")

title = driver.title

login = findByName("login")
login.click()

expTitle = driver.title 


'''----------------------------------ADDING EQUIPMENT--------------------------------'''
findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(4) > a').click() 
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(1) > label').click()
equipmentName = 'Tastatura'
findAndSendByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(1) > div > div:nth-child(1) > form > input[type=text]:nth-child(2)', equipmentName)
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(1) > div > div:nth-child(1) > form > input.button.blue').click()

time.sleep(2)
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(1) > label').click()
time.sleep(2)
searchThroughTableforEquipment(equipmentName)

if cellValue == equipmentName:
    testPassed()
else:
    testFailed()

'''----------------------------------UPDATING AND CHECKING IF EQUIPMENT IS UPDATED-------------------------------------------------'''

findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(4) > a').click()
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(1) > label').click()
searchThroughTableforEquipment(equipmentName)
newEquipmentName = 'Skener'
findByCSSSelector('#myTable > tbody > tr:nth-child('+str(i)+') > td:nth-child(3) > button:nth-child(1)').click()
time.sleep(3)
equipmentFieldName = findByCSSSelector('#modaltype > div > div.modal-body > div > form > input[type=text]:nth-child(1)')
equipmentFieldName.clear()
equipmentFieldName.send_keys(newEquipmentName)
findByCSSSelector('#modaltype > div > div.modal-body > div > form > input.button.blue').click()

time.sleep(5)

findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(1) > label').click()

searchThroughTableforEquipment(newEquipmentName)

if cellValue == newEquipmentName:
    testPassed()
else:
    testFailed()

'''-------------------------------------------------DELETE EQUIPMENT---------------------------------------------------'''

findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(4) > a').click()
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(1) > label').click()

searchThroughTableforEquipment(newEquipmentName)
time.sleep(2)

findByCSSSelector('#myTable > tbody > tr:nth-child('+str(i)+') > td:nth-child(3) > button.button.red').click()
time.sleep(2)
findByCSSSelector('#del').click()

searchThroughTableforEquipment(newEquipmentName)

if cellValue == newEquipmentName:
    testFailed()
else:
    testPassed()

time.sleep(5)

'''----------------------------------------ADDING BRAND------------------------------------------------------'''

findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(4) > a').click()
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > label').click()
brandName = 'ALIENWARE'
findAndSendByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > div > div:nth-child(1) > form > input[type=text]:nth-child(2)', brandName)
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > div > div:nth-child(1) > form > input.button.blue').click()
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > label').click() #click the dropdown again because it closes when submitting 

searchThroughTableforBrand(brandName)

if cellValue == brandName:
    testPassed()
else:
    testFailed()

time.sleep(5)
'''----------------------------------------UPDATING BRAND------------------------------------------------------'''

findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(4) > a').click()
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > label').click()

driver2 = findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > div') #load the table div to the driver variable so we can search through it, the 1st and 2nd table have the same selector

newBrandName = 'Fujitsu'

searchThroughTableforBrand(brandName)

findByCSSSelector2('#myTable > tbody > tr:nth-child('+str(i)+') > td:nth-child(3) > button:nth-child(1)').click()
time.sleep(2)
findByCSSSelector('#modalproducer > div > div.modal-body > div > form > input[type=text]:nth-child(1)').clear()
findAndSendByCSSSelector('#modalproducer > div > div.modal-body > div > form > input[type=text]:nth-child(1)', newBrandName)
findByCSSSelector('#modalproducer > div > div.modal-body > div > form > input.button.blue').click()
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > label').click()

searchThroughTableforBrand(newBrandName)

if cellValue == newBrandName:
    testPassed()
else:
    testFailed()

time.sleep(5)
'''----------------------------------------DELETNG BRAND------------------------------------------------------'''

findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(4) > a').click()
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > label').click()

driver2 = findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > div')

time.sleep(3)
findByCSSSelector2('#myTable > tbody > tr:nth-child('+str(i)+') > td:nth-child(3) > button.button.red').click()
findByCSSSelector('#del').click()
time.sleep(3)
findByCSSSelector('#wrapper > div:nth-child(2) > div:nth-child(2) > label').click()

searchThroughTableforBrand(newBrandName)

if cellValue == newBrandName:
    testFailed()
else:
    testPassed()


driver.quit()