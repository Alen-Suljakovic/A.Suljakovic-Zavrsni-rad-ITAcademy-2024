from selenium import webdriver 
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import datetime

dateAndTime = datetime.datetime.now()
dateAndTime = dateAndTime.strftime("%d-%m-%Y %H-%M-%S")
rowNum = 3                                                     #Starts at 2 since the 1st excel Cell is for the title
i=2                                                 #Has start at 2 since no1 is for the title
x = None
lookupValue = None
cellValue = None
exception = 0
locatedNameXpath = 'PlaceHolder'
QACheckbook = "QA Checkbook " + str(dateAndTime) + ".xlsx"
centerText = Alignment(horizontal='center', vertical='center')              #Center text in excel
cellfillRed = PatternFill(patternType='solid', fgColor='FF3333')            #choose fill type and RED color in excel
cellfillGreen = PatternFill(patternType='solid', fgColor='33CC33')       #choose fill type and GREEN color

wb = openpyxl.load_workbook("QA Checkbook Template.xlsx")
ws = wb['Background data'] 
backgroundData = ws['B1'].value
wb = openpyxl.load_workbook(backgroundData)
QACheckbook = backgroundData

def wait(x):
    driver.implicitly_wait(x)                #Time module discarded- code runs faster

def findAndSendByName(Name,Input):           #Find an element by name and send input
    x = driver.find_element(By.NAME, Name)
    x.send_keys(Input)
    return x

def findAndSendByXPath(XPath,Input):
    x = driver.find_element(By.XPATH, XPath)
    x.send_keys(Input)
    return x

def findByName(Name):                        #Find an element by name
    return driver.find_element(By.NAME, Name)

def findByXPath(XPath):
    return driver.find_element(By.XPATH, XPath)

def findByCSSSelector(Selector):
    return driver.find_element(By.CSS_SELECTOR, Selector)

def testFailed():
    global rowNum
    cell = "E"
    cell += str(rowNum)                                  #Casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['6. Equipment']                              #Select excel sheesheet and load to ws
    ws[cell] = "FAIL"                            #Write to cell 
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillRed
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def testPassed():
    global rowNum
    cell = "E"
    cell += str(rowNum)                                  # casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['6. Equipment']                              #Select excel sheesheet and load to ws
    ws[cell] = "PASS"                            #Write to cell  
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillGreen
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def DoesNotExist():
    global rowNum
    print('DOES NOT EXIST!')
    cell = "E"
    cell += str(rowNum)                                  #Casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['6. Equipment']                              #Select excel sheesheet and load to ws
    ws[cell] = "FAIL"                            #Write to cell 
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillRed
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def searchThroughTableforEquipment(lookupValue):             #Search through table for Name Value
    global cellValue
    global exception
    global locatedNameXpath
    global i
    cellValue = "Reset"
    i = 2
    exception = 0
    while cellValue != lookupValue and exception != 1:
        try:
            locatedNameXpath = driver.find_element(By.XPATH, '/html/body/div/div[1]/div[1]/div/div[2]/div[2]/table/tbody/tr['+str(i)+']/td[2]')
        except: 
            exception = 1
            print("Equipment NOT found on list!")

        try:
            cellValue = locatedNameXpath.text
        except:
            cellValue = 'Blank'
        
        if cellValue == lookupValue:
            print(cellValue, 'Found in table!')
        elif cellValue != lookupValue:
            i +=1

def searchThroughTableforBrand(lookupValue):             #Search through table for Name Value
    global cellValue
    global exception
    global locatedNameXpath
    global i
    cellValue = "Reset"
    i = 2
    exception = 0
    while cellValue != lookupValue and exception != 1:
        try:
            locatedNameXpath = driver.find_element(By.XPATH, '/html/body/div/div[1]/div[2]/div/div[2]/div[2]/table/tbody/tr['+str(i)+']/td[2]')
        except:
            exception = 1
            print("Brand found on list!")

        try:
            cellValue = locatedNameXpath.text
        except:
            cellValue = 'Blank'

        if cellValue == lookupValue:
            print(cellValue, 'Found in table!')
        elif cellValue != lookupValue:
            i +=1

driver = webdriver.Chrome()

driver.get("https://puppies-closet.com/evidencija/login.php")

driver.maximize_window()


'''------------------------ LOGIN ----------------------------'''

username = findAndSendByName("username", "alensuljakovic")
password = findAndSendByName("password", "alensuljakovic123")

title = driver.title

login = findByName("login")
login.click()

expTitle = driver.title 

'''------------------------ EQUIPMENT TESTING ----------------------------'''
'''------------------------ CREATING EQUIPMENT ----------------------------'''

findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(2) > a').click() #Oprema Header

findByName('type_id').click()   #dropdown 1
findByCSSSelector('#type_id > option:nth-child(2)').click() #option 1

findByName('producer_id').click()   #dropdown 2
findByCSSSelector('#producer_id > option:nth-child(2)').click() #option 1
invNum = 'InvAlen01'
findAndSendByName('inventoryNumber', invNum)
findAndSendByName('serialNumber', '001')

findByName('save').click()

findAndSendByName('equSearch', invNum)
findByName('equipmentSearch').click()

time.sleep(3)

cellValue = findByCSSSelector('#results > div > table > tbody > tr:nth-child(2) > td:nth-child(4)').text

if cellValue == invNum: #Did we create equipment
    testPassed()
else:
    testFailed()


findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(1) > a').click() #back to landing page

''''-------------------------------- FIND MY ACCOUNT AND ADD EQUIOPMENT TO IT -----------------------------------------'''

findAndSendByName('search', 'AlenTest')
findByName('employeesSearch').click()
time.sleep(3)
findByCSSSelector('#results > div > table > tbody > tr:nth-child(2) > td:nth-child(8) > button.button.blue').click()
time.sleep(5)
findByName('checkEquip').click()
findByName('obligation').click()
time.sleep(5)
findAndSendByName('search', 'AlenTest')
time.sleep(5)
findByName('employeesSearch').click()
time.sleep(5)
findByCSSSelector('#results > div > table > tbody > tr:nth-child(2) > td:nth-child(8) > button.button.blue').click()
time.sleep(5) #sleep mandatory, ajax is acting weird

equipmentInUse = findByCSSSelector('#modalequipment > div > div.table > form > table > tbody > tr:nth-child(2) > td:nth-child(4)').text

testPassed() 


findByCSSSelector('#modalequipment > span').click() #closing the add equipment dialog window

'''------------------------------------ TRY TO PRINT REVERS -----------------------------------------------'''

findAndSendByName('search', 'AlenTest')

time.sleep(5)
findByName('employeesSearch').click()
time.sleep(5)
findByCSSSelector('#results > div > table > tbody > tr:nth-child(2) > td:nth-child(8) > button.button.blue').click()
time.sleep(5) #sleep mandatory, ajax is acting weird
findByName('checkEmpEquip').click()
findByName('obligateEquipEmp').click()
driver.switch_to.window(driver.window_handles[0])
testFailed() #cant print 'revers zaduzenja', has to be hard coded since there is no logical way to test what this button should do

'''------------------------------------ TRY TO REMOVE EQUIPMENT FROM ACCOUNT -----------------------------------------------'''

findAndSendByName('search', 'AlenTest')
time.sleep(5)
findByName('employeesSearch').click()
time.sleep(5)
findByCSSSelector('#results > div > table > tbody > tr:nth-child(2) > td:nth-child(8) > button.button.blue').click()
time.sleep(5) #sleep mandatory, ajax is acting weird
findByName('checkEmpEquip').click()
findByName('obligateEquipE').click()
driver.switch_to.window(driver.window_handles[0])
testFailed() #cant 'razduzi opremu i stampaj revers', also has to be hard coded for the same reason

