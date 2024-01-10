from selenium import webdriver 
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import datetime

dateAndTime = datetime.datetime.now()
dateAndTime = dateAndTime.strftime("%d-%m-%Y %H-%M-%S")
rowNum = 3                                                     #Starts at 2 since the 1st excel Cell is for the title
i=2                                                 #Has start at 2 since no1 is for the title
x = None
lookupValue = None
cellValue = None
exception = 0
QACheckbook = "QA Checkbook " + str(dateAndTime) + ".xlsx"
centerText = Alignment(horizontal='center', vertical='center')              #Center text in excel
cellfillRed = PatternFill(patternType='solid', fgColor='FF3333')            #choose fill type and RED color in excel
cellfillGreen = PatternFill(patternType='solid', fgColor='33CC33')       #choose fill type and GREEN color

wb = openpyxl.load_workbook("QA Checkbook Template.xlsx")
ws = wb['Background data'] 
backgroundData = ws['B1'].value
ws['B1'] = ""                                                     #Clears the Background data (B2) field in excel so further tests can be ran
wb.save("QA Checkbook Template.xlsx")
wb = openpyxl.load_workbook(backgroundData)
QACheckbook = backgroundData

def wait(x):
    driver.implicitly_wait(x)                #Time module discarded- code runs faster

def findAndSendByName(Name,Input):           #Find an element by name and send input
    x = driver.find_element(By.NAME, Name)
    x.send_keys(Input)
    return x

def findAndSendByXPath(XPath,Input):
    x = driver.find_element(By.XPATH, XPath)
    x.send_keys(Input)
    return x

def findByName(Name):                        #Find an element by name
    return driver.find_element(By.NAME, Name)

def findByXPath(XPath):
    return driver.find_element(By.XPATH, XPath)

def findByCSSSelector(Selector):
    return driver.find_element(By.CSS_SELECTOR, Selector)

def testFailed():
    global rowNum
    cell = "E"
    cell += str(rowNum)                                  #Casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['4. Reports']                              #Select excel sheesheet and load to ws
    ws[cell] = "FAIL"                            #Write to cell 
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillRed
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def testPassed():
    global rowNum
    cell = "E"
    cell += str(rowNum)                                  # casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['4. Reports']                              #Select excel sheesheet and load to ws
    ws[cell] = "PASS"                            #Write to cell  
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillGreen
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

def DoesNotExist():
    global rowNum
    print('DOES NOT EXIST!')
    cell = "E"
    cell += str(rowNum)                                  #Casting x as string and adding it to cell value for ws[cell] use
    rowNum += 1

    ws = wb['4. Reports']                              #Select excel sheesheet and load to ws
    ws[cell] = "FAIL"                            #Write to cell 
    ws[cell].alignment = centerText
    ws[cell].fill = cellfillRed
    wb.save(QACheckbook)
    print("Saved to Workbook:", QACheckbook)

driver = webdriver.Chrome()

driver.get("https://puppies-closet.com/evidencija/login.php")

driver.maximize_window()


'''------------------------ LOGIN ----------------------------'''

username = findAndSendByName("username", "alensuljakovic")
password = findAndSendByName("password", "alensuljakovic123")

title = driver.title

login = findByName("login")
login.click()


'''------------------------------------BLANK EQUIPMENT PER ORGANIZATION REPORT----------------------------------'''

driver.get("https://puppies-closet.com/evidencija/index.php")
time.sleep(5)
findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(3) > a').click() #click on 'Izvjestaji' 
findByName('orgReport').click() #Submit
driver.switch_to.window(driver.window_handles[1])

pageSource = driver.page_source

if 'div' in pageSource:
    print('Not blank')
    testPassed()
else:
    print('blankpage')
    testFailed()

driver.switch_to.window(driver.window_handles[0])
time.sleep(5)

'''------------------------------------BLANK EQUIPMENT PER OFFICE REPORT----------------------------------'''

driver.get("https://puppies-closet.com/evidencija/index.php")
findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(3) > a').click() #click on 'Izvjestaji' 
findByName('officeReport').click() #Submit
driver.switch_to.window(driver.window_handles[2])

pageSource = driver.page_source

if 'div' in pageSource:
    print('Not blank')
    testPassed()
else:
    print('blankpage')
    testFailed()

driver.switch_to.window(driver.window_handles[0])
time.sleep(5)
'''------------------------------------ BLANK EQUIPMENT PER EMPLOYEE REPORT----------------------------------'''

driver.get("https://puppies-closet.com/evidencija/index.php")
findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(3) > a').click() #click on 'Izvjestaji'
findByName('empReport').click() #Submit
driver.switch_to.window(driver.window_handles[3])

pageSource = driver.page_source

if 'div' in pageSource:
    print('Not blank')
    testPassed()
else:
    print('blankpage')
    testFailed()

driver.switch_to.window(driver.window_handles[0])


'''------------------------------------EQUIPMENT PER ORGANIZATION REPORT----------------------------------'''

time.sleep(5)
findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(3) > a').click() #click on 'Izvjestaji' 
findByName('org_id').click() #dropdown
findByCSSSelector('#org_id > option:nth-child(5)').click() #choice
findByName('orgReport').click() #Submit
driver.switch_to.window(driver.window_handles[1])

pageSource = driver.page_source

if 'div' in pageSource:
    print('Not blank')
    testPassed()
else:
    print('blankpage')
    testFailed()

driver.switch_to.window(driver.window_handles[0])
time.sleep(5)

'''------------------------------------EQUIPMENT PER OFFICE REPORT----------------------------------'''

driver.get("https://puppies-closet.com/evidencija/index.php")
findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(3) > a').click() #click on 'Izvjestaji' 
findByName('office_id').click() #dropdown
findByCSSSelector('#office_id > option:nth-child(5)').click() #choice
findByName('officeReport').click() #Submit
driver.switch_to.window(driver.window_handles[2])

pageSource = driver.page_source

if 'div' in pageSource:
    print('Not blank')
    testPassed()
else:
    print('blankpage')
    testFailed()

driver.switch_to.window(driver.window_handles[0])
time.sleep(5)
'''------------------------------------EQUIPMENT PER EMPLOYEE REPORT----------------------------------'''

driver.get("https://puppies-closet.com/evidencija/index.php")
findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(3) > a').click() #click on 'Izvjestaji'
findByName('employees_id').click() #dropdown
findByCSSSelector('#employees_id > option:nth-child(4)').click() #choice
findByName('empReport').click() #Submit
driver.switch_to.window(driver.window_handles[3])

pageSource = driver.page_source

if 'div' in pageSource:
    print('Not blank')
    testPassed()
else:
    print('blankpage')
    testFailed()

driver.switch_to.window(driver.window_handles[0])

'''------------------------------------FREE EQUIPMENT REPORT----------------------------------'''

driver.get("https://puppies-closet.com/evidencija/index.php")
findByCSSSelector('#wrapper > header > nav > ul:nth-child(1) > li:nth-child(3) > a').click() #click on 'Izvjestaji'
findByCSSSelector('#wrapper > div > div.free > a').click() #choice 
driver.switch_to.window(driver.window_handles[4])

pageSource = driver.page_source

if 'div' in pageSource:
    print('Not blank')
    testPassed()
else:
    print('blankpage')
    testFailed()

driver.switch_to.window(driver.window_handles[0])

driver.quit()